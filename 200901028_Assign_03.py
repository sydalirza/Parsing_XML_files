#Syed Ali Raza 200901028 CS-01-B 

import xml.etree.ElementTree as ET
from openpyxl import Workbook
import os

#Reads input file
input_file = "compiler.xml"

#Creates tree to parse 
tree = ET.parse(input_file)
root = tree.getroot()

#List of column names
column_names = ["Book Id", "Author Name", "Title", "Genre", "Price", "Publish Date", "Description"]

#All required details are appended into a list
details = []
for child in root:
    details.append([child.get("id"), child.find("author").text, child.find("title").text, child.find(
        "genre").text, child.find("price").text, child.find("publish_date").text, child.find("description").text])

#Creating Workbook Class object
workbook = Workbook()
spreadsheet = workbook.active

#Writes column names
for j, value in enumerate(column_names):
        spreadsheet.cell(row=1, column=j+1).value = column_names[j]

#Inserts values into spreadsheet
for i, row in enumerate(details):
    for j, value in enumerate(row):
        spreadsheet.cell(row=i+2, column=j+1).value = value

#Creates new file
new_filename = os.path.abspath("./200901028_Assign_03.xlsx")
workbook.save(new_filename)